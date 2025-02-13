# Copyright (c) 2020 - owgs
# License: Open-source MIT

'''
Wrapper functions for the openpyxl library.
'''

# Import dependencies from py standard libs
import re

import os

import pandas as pd

from copy import copy

# Import dependencies from openpyxl
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment
utils = openpyxl.utils

import pyeasylib

# Start of class

class CellRangeIterator:
    '''
    Cell iterator to format cells easily.
    #CHANGES
    #20200422 - initialised by owgs    
    '''
    
    def __init__(self, cell_range, ws):
        
        # Get the anchors
        anchors = cell_range.upper().split(":")
        num_anchors = len(anchors)
        if num_anchors == 1:
            anchor1 = anchors[0]
            anchor2 = anchor1
        elif num_anchors == 2:
            anchor1, anchor2 = anchors
        else:
            raise Exception ("Invalid cell range: %s" % cell_range)
        
        # Split the anchors
        row1, col_alpha_1 = self.__split_anchor__(anchor1)
        row2, col_alpha_2 = self.__split_anchor__(anchor2)
        
        # reorder the rows
        rows= sorted([row1, row2])
        
        # reorder the columns
        col_alphas = sorted([col_alpha_1, col_alpha_2])
        cols = [
            openpyxl.utils.cell.column_index_from_string(col_alphas[0]),
            openpyxl.utils.cell.column_index_from_string(col_alphas[1])
            ]

        # Save to self
        self.ws = ws
        self.cell_range = cell_range
        self.rows = rows
        self.cols = cols
        self.col_alphas = col_alphas
        
        
    def iterate_over_range(self):
        '''
        for cell in self.iterate_over_range():
            
            # do something with cell
            ... ...            
        '''       
        
        # Access the rows and cols
        min_row, max_row = self.rows
        min_col, max_col = self.cols
        
        # Generator
        for ridx in range(min_row, max_row+1):
            
            for cidx in range(min_col, max_col+1):
                
                yield self.ws.cell(ridx, cidx)
                
                
    def apply_cell_function_over_range(self, cell_function):
        '''
        Applies a custom function to modify the range.
        Compare with apply_cell_format_over_range. 
        
        For example:
        
            def cell_function(cell):
                cell.value = 1; 
                cell.font = Font(color="008000")
        
            self = CellRangeIterator("A1:C10", ws)
            self.apply_cell_function_over_range(cell_function)
        '''
        
        for cell in self.iterate_over_range():
            
            cell_function(cell)
    
    
    def apply_cell_format_over_range(self, cell_attribute_dict):
        '''
        Applies formatting to cell based on setting attributes.
        Compare with apply_cell_function_over_range.
        
        For example:
            cell_range3 = "H2:J10"
            self = CellRangeIterator(cell_range3, ws)
            self.apply_cell_format_over_range({"value": 2, 
                                                "font": Font(color='008000')
                                                })
        '''
        
        for cell in self.iterate_over_range():
            
            apply_cell_format(cell, cell_attribute_dict)
            
        
    def __split_anchor__(self, anchor):
        '''
        split anchor = A10 to A and 10
        
        Returns row and column.
        '''
        
        # Split to extract the alpha and digit
        reg_exp = r'\D+|\d+'
        extracted = re.findall(reg_exp, anchor)
        num_extracted = len(extracted)
        
        # Get the row and column for this anchor
        column = None; row = None
        if num_extracted == 2:
            
            column, row = extracted
                    
        elif num_extracted == 1:
            
            if extracted[0].isalpha():
                column = extracted[0]
            else:
                row = extracted[0]
        
        else:
            
            raise Exception ("Unexpected anchor: %s." % anchor)
        
        # Format row to int
        try:
            
            row = int(row)
        
        except ValueError:            
        
            raise Exception ("Unexpected anchor: %s." % anchor)
                        
        return row, column
        
        
def format_cell_to_2_decimals_or_dash(cell):
    '''
    Format to 2 decimals, or to dash if 0.
    '''
    
    cell.number_format = '0.00;-0.00;"-";@'
        
def convert_df_to_excel_rows_cols(df0):
    '''
    assume df0 index and columns starts from 0.
    '''
    
    def __check_0_start_and_contiguous__(index):
        
        # check 0
        if index[0] != 0:
            raise Exception (
                f"df index/columns does not start from 0: {index}"
                )
        
        # Check contiguous
        expected_vals = range(index.shape[0])
        if list(expected_vals) != index.values.tolist():
            raise Exception (
                f"df index/columns does not have contiguous range: {index}"
                )

    # make a copy to avoid changing the original input
    df = df0.copy()
    
    if df.shape[0] > 0:
            
        __check_0_start_and_contiguous__(df.index)
        __check_0_start_and_contiguous__(df.columns)
        
        # Update to xl rows and cols    
        df.index = df.index + 1
        df.index.name = "ExcelRow"
    
        df.columns = (df.columns + 1).map(utils.get_column_letter)
        df.columns.name = "ExcelCol"
    
    return df    
    

def read_excel_with_xl_rows_cols(fp, sheet_name = 0):
    '''
    Read Excel sheets but force rows and cols to Excel alphanum.
    '''
    
    data = pd.read_excel(
        fp, 
        sheet_name = sheet_name,
        index_col = None,   # Hardcoded 
        header = None,      # Hardcoded
        skiprows = 0        # Hardcoded
        )
    
    # CHeck type
    type_data = type(data)
    if type_data is pd.DataFrame:
        output = convert_df_to_excel_rows_cols(data) 
    elif type_data is dict:
        output = {}
        for sheet, df in data.items():
            output[sheet] = convert_df_to_excel_rows_cols(df)
    else:
        raise TypeError (
            f"Unexpected format: {type_data}"
            )
    return output
    
def range_excel_letters(start, stop, include_right=False, step=1):
    """
    Get a range of Excel column alphabets from the start and stop 
    alphabet.
    
    If include_right = True, the stop alpha will be included.
    
    Usage:
        > range_excel_letters("A", "C")
        Out[4]: ['A', 'B']
        
        > range_excel_letters("A", "C", True)
        Out[5]: ['A', 'B', 'C']
        
        > range_excel_letters("A", "C", True, 2)
        Out[6]: ['A', 'C']
        
        > range_excel_letters("A", "C", False, 2)
        Out[7]: ['A']
    """ 
    
    # conditions
    if not start.isalpha():
        raise Exception (f"start is not a letter: {start}.")
        
    if not stop.isalpha():
        raise Exception (f"stop is not a letter: {stop}.")
        
    # add delta, if include right is True
    delta = 1 if include_right else 0
    
    # start_index
    start_idx = utils.column_index_from_string(start)
    stop_idx = utils.column_index_from_string(stop)
    
    # Get the data    
    output_list = []
    for idx in range(start_idx, stop_idx+delta, step):
        output_list.append(utils.get_column_letter(idx))
        
    return output_list


def read_ws(ws, drop_na_rows = False, drop_na_cols = False):
    '''
    Read data from Excel worksheet into pandas dataframe.

    Parameters
    ----------
    ws : openpyxl worksheet
    drop_na_rows : boolean, optional
        If True, NA rows will be dropped. The default is False.
    drop_na_cols : boolean, optional
        If True, NA cols will be dropped. The default is False.

    Returns
    -------
    df
    '''
    
    # Get the boundaries
    max_column = ws.max_column
    max_row = ws.max_row
    
    # Create the df
    df = pd.DataFrame(
        index = pd.Index(range(1, max_row+1), name="ExcelRow"),
        columns = pd.Index(range(1, max_column+1), name="ExcelCol")
        )
    
    # Loop and fill the values
    for c in range(1, max_column+1):
        
        for r in range(1, max_row+1):

            df.at[r, c] = ws.cell(r, c).value
            
    # rename the df.column
    df.columns = df.columns.map(openpyxl.utils.get_column_letter)
    
    # drop NA rows
    if drop_na_rows:
        df = df.dropna(how='all')
    
    # drop NA cols
    if drop_na_cols:
        df = df.dropna(axis=1, how='all')
    
    return df


def get_cell_format(
    cell,
    attrnames = ["value", "font", "fill", "number_format", 
                 "alignment", "border", "comment", 
                 "hyperlink", "number_format", "protection"]):
    
    cell_format = {
        attrname: copy(getattr(cell, attrname))
        for attrname in attrnames
        }
    
    return cell_format

def apply_cell_format(
    cell, 
    cell_format):
    
    for cell_attr, cell_attr_val in cell_format.items():
        setattr(cell, cell_attr, cell_attr_val)
            
def copy_worksheet(source_ws, target_wb, new_ws_title=None):
    '''
    Copy a worksheet to a separate target workbook.
    
    This method was written, as it appears that there is no existing
    method/class to do so when we want to copy to a separate workbook.

    Parameters
    ----------
    source_ws : openpyxl worksheet
    target_wb : openpyxl workbook
    
    Returns:
        - target_ws (after copied)
    '''    
    
    if source_ws.parent == target_wb:
        
        # if the ws's wb and the target wb is the same, there
        # is already a way to do so using inbuilt method.
        target_ws = target_wb.copy_worksheet(source_ws)   
        
    else:
        # Title of the source ws
        source_title = source_ws.title
        
        # Create a new sheet in the target wb
        target_ws = target_wb.create_sheet(source_title+"new")
        
        # Now, will loop through
        source_df = read_ws(source_ws, drop_na_rows=True, drop_na_cols=True)
        
        # format attributes
        #for some reason, if we add 'style', it will reset all the formatting
        attrnames = ["value", "font", "fill", "number_format", "alignment",
                     "border", "comment", "hyperlink", "number_format",
                     "protection"] 
        
        # Now we loop through each of the source data and push it to the target
        for c in source_df.columns:
            
            for r in source_df.index:
                
                # Get the source cell
                loc = f"{c}{r}"
                source = source_ws[loc]
                        
                # Copy all the attributes            
                for attrname in attrnames:
                    attr = getattr(source, attrname)
                    setattr(target_ws[loc], attrname, copy(attr))
                
        # hide the hidden rows
        for r in source_df.index:
            target_ws.row_dimensions[r].hidden = source_ws.row_dimensions[r].hidden
            
        # adjust column width
        for c in source_df.columns:
            target_ws.column_dimensions[c].width = source_ws.column_dimensions[c].width
         
        # adjust row height
        for r in source_df.index:
            target_ws.row_dimensions[r].height = source_ws.row_dimensions[r].height
            
        # Identifying merged cells in the source worksheet
        merged_cell_ranges = source_ws.merged_cells.ranges
        
        # Loop through the merged cell coords and apply to target ws
        for merged_cell_range in merged_cell_ranges:
            
            # Get the coord location
            coord = merged_cell_range.coord #e.g. F9:F11
            
            # merge in the target ws
            target_ws.merge_cells(coord)
    
    # rename the sheetname if required
    if new_ws_title is not None:
        
        target_ws.title = new_ws_title
        
    # Return
    return target_ws

def df_to_worksheet(df, ws, 
                    index=True, header=True,
                    startrow=1, startcol=1
                    ):
    '''
    This function writes a dataframe to an openpyxl worksheet.
    
    This is analogous to the pd.DataFrame.to_excel, which takes
    in a pandas ExcelWriter rather than a worksheet directly.
    
    Inputs:
        - df: A pandas dataframe
        - ws: An openpyxl worksheet object
        - index: bool (default True) - write the index
        - header: bool (default True) - write the header
        - startrow: default (1) - base 1 index where 1 = first row
        - startcol: default (1) - base 1 index where 1 = column A
    
    -----------------------------------------------------------
    # Sample usage
    > wb = openpyxl.Workbook()
    > ws = wb.create_sheet()
    
    > df = pd.DataFrame([[1,2,3], [4,5,6], [7,8,9]], index=list("ABC"),
                      columns=["C1", "C2", "C3"])
        
    > df_to_worksheet(df, ws, index=False, header=False,
                    startrow=3, startcol=3)
    
    >wb.save("test to worksheet.xlsx")
    
    ----------------------------------------------------------
    
    CHANGELOGS:
    20200817 - initialised by owgs
    '''

    # Prepare the dataframe based on whether need to write index or header
    #    
    if index is True:
        df = df.reset_index()
        
    if header is True:
        df = df.T.reset_index().T
    
    # Get the number of rows and columns to write
    num_rows, num_cols = df.shape
    
    # Set the start row and cols
    for r_offset in range(num_rows):
        
        ridx = startrow + r_offset
        
        for c_offset in range(num_cols):
            
            value = df.iat[r_offset, c_offset]
    
            # Get the cell
            cidx = startcol + c_offset
            c_alpha = utils.get_column_letter(cidx)
            cell = "%s%s" % (c_alpha, ridx)
            
            # Set the value
            ws[cell] = value
    
    # end of function
    
def df_to_ws_preserving_template_column_order(
        ws, df, expected_template_headers,
        write_remaining_df_columns=True,
        additional_column_font = Font(italic=True) ,
        additional_column_comment_author = os.getlogin(),
        ):
    """
    Writes a DataFrame to an Excel worksheet while preserving the 
    template column order.
    
    This function reads the specified worksheet, identifies the 
    expected header row, and reorders the input DataFrame (`df`) to 
    match the column order in the worksheet template. Any additional 
    columns in `df` not present in the template are appended to the 
    end. The reordered DataFrame is then written to the worksheet, 
    starting from the row after the identified header.
    
    For consistency, the df index will not be saved to the output.
    If you wish to save the index, please reset the index so that 
    it is in the df itself, before passing it to this method.
    
    Args:
        sheetname (str): Name of the worksheet to write data to.
        df (pd.DataFrame): DataFrame containing data to be written.
        expected_headers (list of str): Expected headers to locate 
            the header row.
        write_remaining_df_columns (bool): True -> write all other columns
                                               even if not in template
                                       False -> do not write if not in
                                                template
    
    Process:
        1. Reads the worksheet into a DataFrame (`wsdf`).
        2. Identifies the header row using `expected_headers`.
        3. Extracts column order from the worksheet template.
        4. Reorders `df` columns and appends extra columns if needed.
        5. Determines the starting row for writing data.
        6. Writes the reordered DataFrame to the worksheet.
    
    Dependencies:
        - `pyeasylib.excellib.read_ws()`: Reads an Excel worksheet.
        - `pyeasylib.pdlib.get_expected_data_row_locations()`: Finds 
          the header row index.
        - `pyeasylib.assert_one_and_get()`: Ensures a single header.
        - `pyeasylib.excellib.df_to_worksheet()`: Writes data to a 
          worksheet.
    
    Returns:
        None (modifies the worksheet in place).
    """

    # Get the worksheet and read the contents
    wsdf = read_ws(ws)
    
    # Get the expected header row
    expected_header_row_locs = pyeasylib.pdlib.get_expected_data_row_locations(
        wsdf, expected_template_headers, return_index=True)
    header_row = pyeasylib.assert_one_and_get(
        expected_header_row_locs, what="header row")
    
    # header in the ws template
    template_header_cols = wsdf.loc[header_row].tolist()
    
    # Sort the input data first by header
    df_col_ordered = df.reindex(columns=template_header_cols, fill_value="")
    
    # Then i fill in the rest of the data from the input
    if write_remaining_df_columns:
        other_cols = [c for c in df.columns if c not in template_header_cols]
        if len(other_cols) > 0:
            df_col_ordered = pd.concat(
                [df_col_ordered, df[other_cols]], axis=1)
            
            # will need to write the header
            # Determine the start column index for new headers
            start_col = len(template_header_cols) + 1  # 1-based index
    
            # Write the headers for the new columns in the worksheet
            for col_idx, col_name in enumerate(other_cols, start=start_col):
                cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                cell.font = additional_column_font
                cell.comment = Comment("Additional column", 
                                       additional_column_comment_author)
                
    # Then get the next row
    startrow = header_row + 1
    
    # Start to write
    df_to_worksheet(
        df_col_ordered, ws, 
        index=False, header=False, startrow=startrow, startcol=1)
    
    return None
    
if __name__ == "__main__":
    
    # TESTER for copy_worksheet
    if False:
        # Test copy worksheet
        fp1 = r"./test/file1.xlsx"
        source_wb = openpyxl.open(fp1)
        source_ws = source_wb["Data1"]
        
        fp2 = r"./test/file2.xlsx"
        target_wb = openpyxl.open(fp2)
        
        # Copy
        copy_worksheet(source_ws, target_wb, 'newname')
        
        # Save
        savefp = "./test/sample_out.xlsx"
        target_wb.save(savefp)
        print (f"Saved to {savefp}.")
        
    # Testing the class for CellRangeIterator
    if False:
        
        # initialise the workbook
        wb = openpyxl.Workbook()
        
        # Get the active sheet
        ws = wb.active
        ws.title = "First sheet"
        
        # For cell range A1:B10
        cell_range = "A1:B10"
        self = CellRangeIterator(cell_range, ws)
        for c in self.iterate_over_range():
            c.value = str(c)
            c.font = Font(color='008000')
        
        # For cell range F2:D90
        cell_range2 = "F2:D90"
        def cell_function(cell):
            cell.value = 1
            cell.font = Font(color="008000")
        self2 = CellRangeIterator(cell_range2, ws)
        self2.apply_cell_function_over_range(cell_function)
        
        # For cell range H2:J10
        cell_range3 = "H2:J10"
        self = CellRangeIterator(cell_range3, ws)
        self.apply_cell_format_over_range({"value": 2, 
                                            "font": Font(color='008000')
                                            })
        #wb.save('test_output.xlsx')
