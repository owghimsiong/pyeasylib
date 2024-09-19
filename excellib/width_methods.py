import os
import openpyxl
from win32com.client import Dispatch

class ColumnsWidthAdjuster:
    '''
    A class to handle Excel file operations using win32com.client.

    Attributes:
        excelfp (str): The file path to the Excel file.
        
    REFERENCES:
    # https://stackoverflow.com/questions/62505403/using-python-win32com-to-get-list-of-excel-worksheets#:~:text=To%20get%20the%20name%20of%20each%20sheet%2C%20you,each%20sheet%2C%20you%20must%20use%20the%20.Name%20method
    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # CHATGPT for docstring
    
    #CHANGE LOGS
    #20230322 - Initialised.
    '''
    
    def __init__(self, excelfp):
        '''
        Initializes an instance of the ExcelHandler class.

        Parameters:
            excelfp (str): The file path to the Excel file.
        '''
        
        self.excelfp = excelfp

    def main(self, sheetnames = None, method = None):

        sheetnames = sheetnames

        if method == None:

            self.autofit_win32com(sheetnames)

        elif method == 'openpyxl':

            self.autofit_openpyxl(sheetnames)


    
    def autofit_win32com(self, sheetnames=None):
        '''
        Performs autofit operations on the specified sheets of the Excel file.

        Parameters:
            sheetnames (list or bool, optional): A list of sheet names to autofit. 
                If True, autofits all sheets. If None or False, no operation is performed. 
                Defaults to None.
        '''
        if not sheetnames:
            print("No sheet names provided. No need.")
            return
        
        try:
            excel = Dispatch('Excel.Application')
            excel.Visible = False  # Set Excel to run in the background
            excel.DisplayAlerts = False
            
            # win 32 only works with absolute path
            self.excelfp_abs = os.path.abspath(self.excelfp)
            
            print(f"Opening Excel file: {self.excelfp_abs}...")

            try:
                wb = excel.Workbooks.Open(self.excelfp_abs)

            except Exception as e:
                print(e)
                
            
            if isinstance(sheetnames, list):
                sn_list = sheetnames
            elif sheetnames is True:
                sn_list = [ws.Name for ws in wb.Sheets]
            else:
                raise ValueError(f"Invalid type for sheetnames: {type(sheetnames)}")
            
            for sheetname in sn_list:
                try:
                    ws = excel.Worksheets(sheetname)
                    ws.Activate()
                    ws.Columns.AutoFit()
                    ws.Rows.AutoFit()
                    print(f"AutoFit done for sheet '{sheetname}'.")
                except Exception as e:
                    print(f"Error while processing sheet '{sheetname}': {e}")
            
            # wb.Save()

            # create new filepath to save it in
            #   temp solution until we figure out why the file cannot be overwritten
            dirname, filename = os.path.split(self.excelfp_abs)
            # Split the file name into the name and extension
            name, ext = os.path.splitext(filename)
            # Create the new file name by adding the suffix
            new_filename = f"{name}_formatted{ext}"
            # Combine the directory and new file name to get the new file path
            new_filepath = os.path.join(dirname, new_filename)

            wb.SaveAs(new_filepath)
            wb.Close()
            excel.Quit()
            # print(f"Excel file saved and closed successfully at: {self.excelfp_abs}.")
            # print(f"Excel file saved and closed successfully at: {new_filepath}.")
            
        except Exception as e:
            print(f"An error occurred: {e}")
            # wb.Close()
            # excel.Quit()
            # print("Closed the file successfully")


    def autofit_openpyxl(self, sheetnames = None):
        wb = openpyxl.load_workbook(self.excelfp)

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]

            max_column = ws.max_column
            max_row = ws.max_row
            
            #assert False
            for col_index in range(1, max_column+1):
                
                col_alpha = openpyxl.utils.get_column_letter(col_index)

                max_width = 0
                for row_index in range(1, max_row+1):
                    
                    value = ws.cell(row_index, col_index).value
                    value2 = "" if value is None else value
                    
                    
                    width = len(str(value2))

                    
                    # update width
                    max_width = max(max_width, width)
                    
                
                                
                    print (col_alpha, col_index, row_index, value, value2, "|", width, max_width)
                    
                # adjust
                adjusted_width = (max_width + 2) * 1.2
                if adjusted_width < 80:
                    ws.column_dimensions[col_alpha].width = adjusted_width
                else:
                    ws.column_dimensions[col_alpha].width = 80
                
                print (col_alpha, adjusted_width)
                
                print ("-" * 50)
            
    
        wb.save(self.excelfp)

if __name__ == "__main__":
    
    
    # Testing adjustment for column width
    if False:
        
        fp = r"./test/file3 - adjust width.xlsx"
        copied_fp = r"./test/file3 - adjust width output.xlsx"

        # Make a copy of the file
        wb = openpyxl.open(fp)
        wb.save(copied_fp)

        # Adjust width on the copied file
        self = ColumnsWidthAdjuster(copied_fp)
        self.main(["Data2"])
    
        assert False, "End of script."
    
    # Non win32 method - KIV only
    if True:
        
        fp = r"./test/file3 - adjust width.xlsx"
        copied_fp = r"./test/file3 - adjust width output.xlsx"

        # Make a copy of the file
        wb = openpyxl.open(fp)
        wb.save(copied_fp)
        
        self = ColumnsWidthAdjuster(copied_fp)
        self.main(["Data2"], method = "openpyxl")